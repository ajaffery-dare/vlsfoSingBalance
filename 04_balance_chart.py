"""
VLSFO Singapore Balance Model
Script 04: Balance Forecast Visualisation

HOW TO RUN:
    streamlit run scripts/04_balance_chart.py

Reads directly from your Excel file — no need to run scripts 01/02 first.
Just update balance.xlsx and run this script.

WHAT IT SHOWS:
  - Total stock with historical mean and ±1 std bands
  - Onshore vs floating stock split
  - Net flow indicator (supply/demand signal)
  - Gross flows: imports, exports, bunker sales, crude imports
  - Days of cover (stock relative to bunker demand)
  - Import origin breakdown (top countries)
  - Export destination breakdown (top countries)
  - Year-on-year comparison of key series
  - Forecast assumptions panel
"""

import streamlit as st
import pandas as pd
import numpy as np
import plotly.graph_objects as go
from plotly.subplots import make_subplots
import openpyxl
from pathlib import Path
from datetime import datetime

# ── PAGE CONFIG ────────────────────────────────────────────────────────────────
st.set_page_config(
    page_title="VLSFO Balance — Singapore",
    page_icon="⚖",
    layout="wide",
    initial_sidebar_state="expanded",
)

st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=IBM+Plex+Mono:wght@400;500;600&family=IBM+Plex+Sans:wght@300;400;500;600&display=swap');
html, body, [class*="css"] { font-family: 'IBM Plex Sans', sans-serif; }
.stApp { background-color: #0a0c10; }
.main  { background-color: #0a0c10; }
.section-hdr {
    font-family: 'IBM Plex Mono', monospace;
    font-size: 10px; letter-spacing: 0.16em; text-transform: uppercase;
    color: #3a4060; border-bottom: 1px solid #1a1e2a;
    padding-bottom: 6px; margin-bottom: 14px; margin-top: 6px;
}
.kpi-card {
    background: #13151c; border: 1px solid #1e2230;
    border-radius: 6px; padding: 14px 16px;
    font-family: 'IBM Plex Mono', monospace;
}
.kpi-label { font-size: 9px; letter-spacing: 0.12em; text-transform: uppercase; color: #3a4060; margin-bottom: 5px; }
.kpi-value { font-size: 22px; font-weight: 600; color: #e8eaf0; line-height: 1; }
.kpi-sub   { font-size: 10px; color: #4a5168; margin-top: 4px; }
.tight  { color: #e05555; }
.loose  { color: #3dba65; }
.neutral{ color: #f59e42; }
div[data-testid="stMetric"] { background:#13151c; border:1px solid #1e2230; border-radius:6px; padding:12px 14px; }
div[data-testid="stMetricLabel"] p { color:#3a4060 !important; font-size:10px !important; letter-spacing:0.1em; text-transform:uppercase; font-family:'IBM Plex Mono',monospace !important; }
div[data-testid="stMetricValue"] { font-family:'IBM Plex Mono',monospace !important; color:#e8eaf0 !important; }
h1,h2,h3 { font-family:'IBM Plex Sans',sans-serif !important; color:#e8eaf0 !important; }
p { color:#8890a8 !important; }
.stTabs [data-baseweb="tab"] { font-family:'IBM Plex Mono',monospace; font-size:11px; letter-spacing:0.06em; color:#4a5168; }
.stTabs [aria-selected="true"] { color:#e8eaf0 !important; }
</style>
""", unsafe_allow_html=True)

# ── CONSTANTS ──────────────────────────────────────────────────────────────────
EXCEL_FILE = Path("data/raw/balance.xlsx")

CHART = dict(
    paper_bgcolor="#0a0c10",
    plot_bgcolor="#0a0c10",
    font=dict(family="IBM Plex Mono, monospace", size=11, color="#4a5168"),
    margin=dict(l=55, r=20, t=35, b=40),
    hovermode="x unified",
)
AX = dict(gridcolor="#161820", showgrid=True, zeroline=False, tickfont=dict(size=10), linecolor="#1a1e2a")

C = dict(
    stock    = "#38bdf8",
    onshore  = "#4f8ef7",
    floating = "#a78bfa",
    imports  = "#3dba65",
    exports  = "#e05555",
    bunker   = "#f59e42",
    crude    = "#fb923c",
    prod     = "#94a3b8",
    netflow_neg = "#e05555",
    netflow_pos = "#3dba65",
    mean     = "#f59e42",
    band     = "rgba(245,158,66,0.08)",
    forecast = "rgba(79,142,247,0.12)",
    act_line = "#38bdf8",
    fct_line = "#4f8ef7",
)

# ── DATA LOADING ───────────────────────────────────────────────────────────────
@st.cache_data(ttl=300)
def load_balance():
    if not EXCEL_FILE.exists():
        return None, None, None, None

    wb = openpyxl.load_workbook(EXCEL_FILE, read_only=True, data_only=True)
    ws = wb["Balance Forecast"]
    raw = [r for r in ws.iter_rows(values_only=True) if any(v is not None for v in r)]

    dates = [v for v in raw[0][1:] if isinstance(v, datetime)]
    n = len(dates)

    def s(i):
        return pd.to_numeric(list(raw[i][1:n+1]), errors="coerce")

    df = pd.DataFrame({"date": pd.to_datetime(dates)})
    df["onshore_kbbl"]   = s(1)
    df["onshore_kt"]     = s(2)
    df["floating_kt"]    = s(3)
    df["total_stock_kt"] = s(4)
    df["stock_change"]   = s(5)
    df["imports"]        = s(6)
    df["exports"]        = s(7)
    df["net_imports"]    = s(8)
    df["bunker_sales"]   = s(9)
    df["crude_imports"]  = s(10)
    df["production"]     = s(11)
    df["imp_vs_bunker"]  = s(12)
    df["netflow"]        = s(19)

    # Derived
    df["days_cover"]     = df["total_stock_kt"] / (df["bunker_sales"] / 30)
    df["supply_ratio"]   = df["net_imports"] / df["bunker_sales"]
    df["gross_supply"]   = df["imports"] + df["crude_imports"] + df["production"]

    # Identify actuals vs forecast
    # Forecast months have floating stock stepping in round increments toward 1400 target
    forecast_floats = {1537.274, 1508.5, 1486.8, 1465.1, 1443.4, 1421.7, 1400.0}
    df["is_forecast"] = (
        df["floating_kt"].round(1).isin([round(v, 1) for v in forecast_floats])
        & (df["date"] >= "2025-10-01")
    )
    df.loc[df["date"] < "2025-10-01", "is_forecast"] = False

    # Historical stats (actuals only, excluding recent 6m to avoid recency bias)
    hist = df[~df["is_forecast"] & (df["date"] < df[~df["is_forecast"]]["date"].max() - pd.DateOffset(months=6))]
    stats = {
        "stock_mean": hist["total_stock_kt"].mean(),
        "stock_std":  hist["total_stock_kt"].std(),
        "netflow_mean": hist["netflow"].mean(),
        "bunker_mean":  hist["bunker_sales"].mean(),
        "imports_mean": hist["imports"].mean(),
        "days_cover_mean": hist["days_cover"].mean(),
    }

    # Read import origins
    ws_imp = wb["Singapore Strait monthly import"]
    imp_raw = [r for r in ws_imp.iter_rows(values_only=True) if any(v is not None for v in r)]
    imp_headers = [h if h is not None else f"col_{i}" for i, h in enumerate(imp_raw[0])]
    imp_data = [r for r in imp_raw[1:] if isinstance(r[0], str) and len(str(r[0])) == 7]
    df_imp = pd.DataFrame(imp_data, columns=imp_headers[:len(imp_data[0])])
    df_imp["date"] = pd.to_datetime(df_imp["date"] + "-01")
    num_cols_imp = [c for c in df_imp.columns if c != "date"]
    df_imp[num_cols_imp] = df_imp[num_cols_imp].apply(lambda x: pd.to_numeric(x, errors="coerce"))

    # Read export destinations
    ws_exp = wb["Singapore Strait exports"]
    exp_raw = [r for r in ws_exp.iter_rows(values_only=True) if any(v is not None for v in r)]
    exp_headers = [h if h is not None else f"col_{i}" for i, h in enumerate(exp_raw[0])]
    exp_data = [r for r in exp_raw[1:] if isinstance(r[0], str) and len(str(r[0])) == 7]
    df_exp = pd.DataFrame(exp_data, columns=exp_headers[:len(exp_data[0])])
    df_exp["date"] = pd.to_datetime(df_exp["date"] + "-01")
    num_cols_exp = [c for c in df_exp.columns if c != "date"]
    df_exp[num_cols_exp] = df_exp[num_cols_exp].apply(lambda x: pd.to_numeric(x, errors="coerce"))

    return df, stats, df_imp, df_exp


# ── CHART HELPERS ──────────────────────────────────────────────────────────────
def split(df, col):
    """Return actual and forecast series with a one-point overlap for continuity."""
    act = df[~df["is_forecast"]][["date", col]].copy()
    fct = df[df["is_forecast"]][["date", col]].copy()
    if len(fct) > 0 and len(act) > 0:
        fct = pd.concat([act.tail(1), fct], ignore_index=True)
    return act, fct


def forecast_band(df):
    """Return x-coordinates for the forecast shading rectangle."""
    fct = df[df["is_forecast"]]
    if len(fct) == 0:
        return None, None
    return fct["date"].iloc[0], fct["date"].iloc[-1]


# ── CHART 1: TOTAL STOCK ───────────────────────────────────────────────────────
def chart_total_stock(df, stats):
    act, fct = split(df, "total_stock_kt")
    act_on, fct_on = split(df, "onshore_kt")
    act_fl, fct_fl = split(df, "floating_kt")

    fig = make_subplots(rows=2, cols=1, row_heights=[0.65, 0.35],
                        shared_xaxes=True, vertical_spacing=0.04)

    # Forecast shading
    fx0, fx1 = forecast_band(df)
    if fx0:
        for row in [1, 2]:
            fig.add_vrect(x0=fx0, x1=fx1, fillcolor=C["forecast"],
                          layer="below", line_width=0, row=row, col=1)

    # Mean and ±1 std bands
    m, sd = stats["stock_mean"], stats["stock_std"]
    fig.add_hrect(y0=m-sd, y1=m+sd, fillcolor=C["band"], layer="below", line_width=0, row=1, col=1)
    fig.add_hline(y=m, line=dict(color=C["mean"], width=1, dash="dot"), row=1, col=1)
    fig.add_annotation(x=df["date"].iloc[-1], y=m, text=f"hist avg {m:.0f}",
                       font=dict(size=9, color=C["mean"]), showarrow=False,
                       xanchor="right", yanchor="bottom", row=1, col=1)

    # Total stock — actual
    fig.add_trace(go.Scatter(
        x=act["date"], y=act["total_stock_kt"], name="Total stock (actual)",
        line=dict(color=C["stock"], width=2),
        hovertemplate="%{x|%b %Y}<br>Total: %{y:.0f} kt<extra></extra>",
    ), row=1, col=1)
    # Total stock — forecast
    if len(fct) > 0:
        fig.add_trace(go.Scatter(
            x=fct["date"], y=fct["total_stock_kt"], name="Total stock (forecast)",
            line=dict(color=C["fct_line"], width=1.5, dash="dash"),
            hovertemplate="%{x|%b %Y}<br>Forecast: %{y:.0f} kt<extra></extra>",
        ), row=1, col=1)

    # Onshore + Floating stacked area (actual only)
    fig.add_trace(go.Scatter(
        x=act_on["date"], y=act_on["onshore_kt"], name="Onshore (kt)",
        stackgroup="stock", line=dict(color=C["onshore"], width=0),
        fillcolor="rgba(79,142,247,0.35)",
        hovertemplate="%{x|%b %Y}<br>Onshore: %{y:.0f} kt<extra></extra>",
    ), row=2, col=1)
    fig.add_trace(go.Scatter(
        x=act_fl["date"], y=act_fl["floating_kt"], name="Floating (kt)",
        stackgroup="stock", line=dict(color=C["floating"], width=0),
        fillcolor="rgba(167,139,250,0.35)",
        hovertemplate="%{x|%b %Y}<br>Floating: %{y:.0f} kt<extra></extra>",
    ), row=2, col=1)
    # Forecast floating
    if len(fct_fl) > 0:
        fig.add_trace(go.Scatter(
            x=fct_fl["date"], y=fct_fl["floating_kt"], name="Floating (fcst)",
            line=dict(color=C["floating"], width=1, dash="dot"), showlegend=False,
        ), row=2, col=1)

    fig.update_layout(**CHART, height=480,
                      legend=dict(orientation="h", y=1.02, x=0, font=dict(size=10), bgcolor="rgba(0,0,0,0)"))
    fig.update_xaxes(**AX)
    fig.update_yaxes(**AX, title_text="Total stock (kt)", row=1, col=1, title_font=dict(size=10))
    fig.update_yaxes(**AX, title_text="Stock split (kt)", row=2, col=1, title_font=dict(size=10))
    return fig


# ── CHART 2: NET FLOW ─────────────────────────────────────────────────────────
def chart_netflow(df, stats):
    act = df[~df["is_forecast"]]
    fct = df[df["is_forecast"]]
    m = stats["netflow_mean"]

    fig = make_subplots(rows=2, cols=1, row_heights=[0.55, 0.45],
                        shared_xaxes=True, vertical_spacing=0.04)

    fx0, fx1 = forecast_band(df)
    if fx0:
        for row in [1, 2]:
            fig.add_vrect(x0=fx0, x1=fx1, fillcolor=C["forecast"],
                          layer="below", line_width=0, row=row, col=1)

    # Historical mean line
    fig.add_hline(y=m, line=dict(color=C["mean"], width=0.8, dash="dot"), row=1, col=1)
    fig.add_hline(y=0, line=dict(color="#2a2f42", width=0.8), row=1, col=1)

    # Net flow bars
    act_colors = [C["netflow_pos"] if v >= 0 else C["netflow_neg"]
                  for v in act["netflow"].fillna(0)]
    fct_colors = [C["netflow_pos"] if v >= 0 else C["netflow_neg"]
                  for v in fct["netflow"].fillna(0)]

    fig.add_trace(go.Bar(
        x=act["date"], y=act["netflow"],
        name="Net flow (actual)", marker_color=act_colors, opacity=0.85,
        hovertemplate="%{x|%b %Y}<br>Net flow: %{y:.0f} kt<extra></extra>",
    ), row=1, col=1)
    if len(fct) > 0:
        fig.add_trace(go.Bar(
            x=fct["date"], y=fct["netflow"],
            name="Net flow (forecast)", marker_color=fct_colors, opacity=0.4,
            hovertemplate="%{x|%b %Y}<br>Forecast: %{y:.0f} kt<extra></extra>",
        ), row=1, col=1)

    # Stock change bars (bottom panel)
    sc_act = act.dropna(subset=["stock_change"])
    sc_colors = [C["netflow_pos"] if v >= 0 else C["netflow_neg"]
                 for v in sc_act["stock_change"]]
    fig.add_hline(y=0, line=dict(color="#2a2f42", width=0.8), row=2, col=1)
    fig.add_trace(go.Bar(
        x=sc_act["date"], y=sc_act["stock_change"],
        name="Stock change MoM", marker_color=sc_colors, opacity=0.75,
        hovertemplate="%{x|%b %Y}<br>MoM change: %{y:.0f} kt<extra></extra>",
    ), row=2, col=1)

    fig.update_layout(**CHART, height=440, barmode="relative",
                      legend=dict(orientation="h", y=1.02, x=0, font=dict(size=10), bgcolor="rgba(0,0,0,0)"))
    fig.update_xaxes(**AX)
    fig.update_yaxes(**AX, title_text="Net flow (kt)", row=1, col=1, title_font=dict(size=10))
    fig.update_yaxes(**AX, title_text="MoM stock Δ (kt)", row=2, col=1, title_font=dict(size=10))
    return fig


# ── CHART 3: GROSS FLOWS ──────────────────────────────────────────────────────
def chart_gross_flows(df):
    act = df[~df["is_forecast"]]
    fct = df[df["is_forecast"]]

    fig = make_subplots(rows=2, cols=1, row_heights=[0.5, 0.5],
                        shared_xaxes=True, vertical_spacing=0.04)

    fx0, fx1 = forecast_band(df)
    if fx0:
        for row in [1, 2]:
            fig.add_vrect(x0=fx0, x1=fx1, fillcolor=C["forecast"],
                          layer="below", line_width=0, row=row, col=1)

    # Top: Imports vs Bunker Sales
    fig.add_trace(go.Scatter(
        x=act["date"], y=act["imports"], name="Imports",
        line=dict(color=C["imports"], width=1.5),
        hovertemplate="%{x|%b %Y}<br>Imports: %{y:.0f} kt<extra></extra>",
    ), row=1, col=1)
    if len(fct) > 0:
        imp_fct = pd.concat([act[["date","imports"]].tail(1), fct[["date","imports"]]])
        fig.add_trace(go.Scatter(
            x=imp_fct["date"], y=imp_fct["imports"],
            line=dict(color=C["imports"], width=1, dash="dash"), showlegend=False,
        ), row=1, col=1)

    fig.add_trace(go.Scatter(
        x=act["date"], y=act["bunker_sales"], name="Bunker sales",
        line=dict(color=C["bunker"], width=1.5),
        hovertemplate="%{x|%b %Y}<br>Bunker: %{y:.0f} kt<extra></extra>",
    ), row=1, col=1)
    if len(fct) > 0:
        bunk_fct = pd.concat([act[["date","bunker_sales"]].tail(1), fct[["date","bunker_sales"]]])
        fig.add_trace(go.Scatter(
            x=bunk_fct["date"], y=bunk_fct["bunker_sales"],
            line=dict(color=C["bunker"], width=1, dash="dash"), showlegend=False,
        ), row=1, col=1)

    fig.add_trace(go.Scatter(
        x=act["date"], y=act["exports"], name="Exports",
        line=dict(color=C["exports"], width=1.5),
        hovertemplate="%{x|%b %Y}<br>Exports: %{y:.0f} kt<extra></extra>",
    ), row=1, col=1)

    # Imports vs Bunker gap (fill)
    fig.add_trace(go.Scatter(
        x=act["date"], y=act["imp_vs_bunker"], name="Imports − bunker",
        line=dict(color="#94a3b8", width=1),
        fill="tozeroy", fillcolor="rgba(148,163,184,0.07)",
        hovertemplate="%{x|%b %Y}<br>Gap: %{y:.0f} kt<extra></extra>",
    ), row=1, col=1)
    fig.add_hline(y=0, line=dict(color="#2a2f42", width=0.8), row=1, col=1)

    # Bottom: Crude imports + Production
    fig.add_trace(go.Bar(
        x=act["date"], y=act["crude_imports"],
        name="Crude imports", marker_color=C["crude"], opacity=0.75,
        hovertemplate="%{x|%b %Y}<br>Crude: %{y:.0f} kt<extra></extra>",
    ), row=2, col=1)
    fig.add_trace(go.Bar(
        x=act["date"], y=act["production"],
        name="Production", marker_color=C["prod"], opacity=0.75,
        hovertemplate="%{x|%b %Y}<br>Production: %{y:.0f} kt<extra></extra>",
    ), row=2, col=1)

    fig.update_layout(**CHART, height=460, barmode="stack",
                      legend=dict(orientation="h", y=1.02, x=0, font=dict(size=10), bgcolor="rgba(0,0,0,0)"))
    fig.update_xaxes(**AX)
    fig.update_yaxes(**AX, title_text="Volume (kt)", row=1, col=1, title_font=dict(size=10))
    fig.update_yaxes(**AX, title_text="Crude + production (kt)", row=2, col=1, title_font=dict(size=10))
    return fig


# ── CHART 4: DAYS COVER ───────────────────────────────────────────────────────
def chart_days_cover(df, stats):
    act, fct = split(df, "days_cover")
    mean_dc = stats["days_cover_mean"]

    fig = go.Figure()

    fx0, fx1 = forecast_band(df)
    if fx0:
        fig.add_vrect(x0=fx0, x1=fx1, fillcolor=C["forecast"],
                      layer="below", line_width=0)

    fig.add_hrect(y0=mean_dc*0.9, y1=mean_dc*1.1,
                  fillcolor=C["band"], layer="below", line_width=0)
    fig.add_hline(y=mean_dc, line=dict(color=C["mean"], width=1, dash="dot"))
    fig.add_annotation(x=df["date"].iloc[-1], y=mean_dc,
                       text=f"hist avg {mean_dc:.1f}d",
                       font=dict(size=9, color=C["mean"]),
                       showarrow=False, xanchor="right", yanchor="bottom")

    fig.add_trace(go.Scatter(
        x=act["date"], y=act["days_cover"],
        name="Days cover", line=dict(color=C["stock"], width=2),
        fill="tozeroy", fillcolor="rgba(56,189,248,0.08)",
        hovertemplate="%{x|%b %Y}<br>Days cover: %{y:.1f}d<extra></extra>",
    ))
    if len(fct) > 0:
        fig.add_trace(go.Scatter(
            x=fct["date"], y=fct["days_cover"],
            name="Days cover (forecast)",
            line=dict(color=C["fct_line"], width=1.5, dash="dash"),
            hovertemplate="%{x|%b %Y}<br>Forecast: %{y:.1f}d<extra></extra>",
        ))

    fig.update_layout(**CHART, height=300,
                      yaxis=dict(**AX, title="Days of cover", title_font=dict(size=10)),
                      xaxis=dict(**AX),
                      legend=dict(orientation="h", y=1.04, x=0, font=dict(size=10), bgcolor="rgba(0,0,0,0)"))
    return fig


# ── CHART 5: IMPORT ORIGINS ───────────────────────────────────────────────────
def chart_import_origins(df_imp, months=12):
    # Top countries by volume over selected period
    df = df_imp.tail(months).copy()
    country_cols = [c for c in df.columns
                    if c not in ("date", "Total", "Ad-hoc", "date")
                    and isinstance(c, str) and df[c].sum() > 0]

    totals = df[country_cols].sum().sort_values(ascending=False)
    top = totals.head(8).index.tolist()
    other_cols = [c for c in country_cols if c not in top]

    palette = ["#4f8ef7","#3dba65","#f59e42","#a78bfa",
               "#38bdf8","#fb923c","#e05555","#94a3b8"]

    fig = go.Figure()
    for i, country in enumerate(top):
        fig.add_trace(go.Bar(
            x=df["date"], y=df[country],
            name=country, marker_color=palette[i % len(palette)], opacity=0.85,
            hovertemplate=f"%{{x|%b %Y}}<br>{country}: %{{y:.0f}} kt<extra></extra>",
        ))
    if other_cols:
        fig.add_trace(go.Bar(
            x=df["date"], y=df[other_cols].sum(axis=1),
            name="Other", marker_color="#2a2f42", opacity=0.7,
            hovertemplate="%{x|%b %Y}<br>Other: %{y:.0f} kt<extra></extra>",
        ))

    fig.update_layout(**CHART, height=320, barmode="stack",
                      legend=dict(orientation="h", y=1.04, x=0, font=dict(size=10), bgcolor="rgba(0,0,0,0)"),
                      yaxis=dict(**AX, title="Imports (kt)", title_font=dict(size=10)),
                      xaxis=dict(**AX))
    return fig


# ── CHART 6: EXPORT DESTINATIONS ─────────────────────────────────────────────
def chart_export_destinations(df_exp, months=12):
    df = df_exp.tail(months).copy()
    country_cols = [c for c in df.columns
                    if c not in ("date", "Total", "ad-hoc")
                    and isinstance(c, str) and df[c].sum() > 0]

    totals = df[country_cols].sum().sort_values(ascending=False)
    top = totals.head(8).index.tolist()
    other_cols = [c for c in country_cols if c not in top]

    palette = ["#e05555","#f59e42","#a78bfa","#4f8ef7",
               "#3dba65","#38bdf8","#fb923c","#94a3b8"]

    fig = go.Figure()
    for i, country in enumerate(top):
        fig.add_trace(go.Bar(
            x=df["date"], y=df[country],
            name=country, marker_color=palette[i % len(palette)], opacity=0.85,
            hovertemplate=f"%{{x|%b %Y}}<br>{country}: %{{y:.0f}} kt<extra></extra>",
        ))
    if other_cols:
        fig.add_trace(go.Bar(
            x=df["date"], y=df[other_cols].sum(axis=1),
            name="Other", marker_color="#2a2f42", opacity=0.7,
            hovertemplate="%{x|%b %Y}<br>Other: %{y:.0f} kt<extra></extra>",
        ))

    fig.update_layout(**CHART, height=320, barmode="stack",
                      legend=dict(orientation="h", y=1.04, x=0, font=dict(size=10), bgcolor="rgba(0,0,0,0)"),
                      yaxis=dict(**AX, title="Exports (kt)", title_font=dict(size=10)),
                      xaxis=dict(**AX))
    return fig


# ── CHART 7: YEAR-ON-YEAR ─────────────────────────────────────────────────────
def chart_yoy(df, col, label, color):
    df = df[~df["is_forecast"]].copy()
    df["year"]  = df["date"].dt.year
    df["month"] = df["date"].dt.month

    years = sorted(df["year"].unique())
    month_labels = ["Jan","Feb","Mar","Apr","May","Jun",
                    "Jul","Aug","Sep","Oct","Nov","Dec"]
    palette_yoy = ["#94a3b8","#4f8ef7","#a78bfa","#3dba65","#f59e42","#38bdf8"]

    fig = go.Figure()
    for i, yr in enumerate(years):
        yr_data = df[df["year"] == yr].sort_values("month")
        lw = 2.0 if yr == years[-1] else 1.0
        opacity = 1.0 if yr == years[-1] else 0.5
        fig.add_trace(go.Scatter(
            x=yr_data["month"], y=yr_data[col],
            name=str(yr), mode="lines+markers",
            line=dict(color=palette_yoy[i % len(palette_yoy)], width=lw),
            opacity=opacity,
            marker=dict(size=4),
            hovertemplate=f"%{{x}}<br>{yr}: %{{y:.0f}}<extra></extra>",
        ))

    fig.update_layout(**CHART, height=300,
                      xaxis=dict(**AX, tickvals=list(range(1,13)), ticktext=month_labels),
                      yaxis=dict(**AX, title=label, title_font=dict(size=10)),
                      legend=dict(orientation="h", y=1.04, x=0, font=dict(size=10), bgcolor="rgba(0,0,0,0)"))
    return fig


# ── MAIN ──────────────────────────────────────────────────────────────────────
def main():
    df, stats, df_imp, df_exp = load_balance()

    if df is None:
        st.error(f"Cannot find: {EXCEL_FILE.resolve()}\nSave your Excel file to data/raw/balance.xlsx")
        return

    actuals  = df[~df["is_forecast"]]
    forecast = df[df["is_forecast"]]
    latest   = actuals.iloc[-1]
    latest_d = latest["date"]

    # ── HEADER ────────────────────────────────────────────────────────────────
    col_t, col_r = st.columns([5, 1])
    with col_t:
        st.markdown(
            "<h1 style='font-family:IBM Plex Mono,monospace;font-size:20px;font-weight:600;"
            "color:#e8eaf0;margin-bottom:2px;letter-spacing:0.04em'>"
            "SINGAPORE VLSFO — SUPPLY & DEMAND BALANCE</h1>"
            "<p style='font-family:IBM Plex Mono,monospace;font-size:10px;"
            "letter-spacing:0.12em;color:#3a4060;margin:0'>"
            "BALANCE FORECAST — MONTHLY</p>",
            unsafe_allow_html=True,
        )
    with col_r:
        if st.button("⟳ Refresh", use_container_width=True):
            st.cache_data.clear()
            st.rerun()

    st.markdown(
        f"<div style='background:#070810;border:1px solid #1a1e2a;border-radius:4px;"
        f"padding:8px 14px;font-family:IBM Plex Mono,monospace;font-size:10px;"
        f"color:#3a4060;margin-bottom:18px'>"
        f"ACTUALS THROUGH: {latest_d.strftime('%b %Y').upper()} &nbsp;│&nbsp; "
        f"FORECAST: {forecast['date'].iloc[0].strftime('%b %Y') if len(forecast) else '—'} "
        f"– {forecast['date'].iloc[-1].strftime('%b %Y') if len(forecast) else '—'} "
        f"&nbsp;│&nbsp; TOTAL HISTORY: {actuals['date'].iloc[0].strftime('%b %Y')} "
        f"– {latest_d.strftime('%b %Y')}</div>",
        unsafe_allow_html=True,
    )

    # ── SIDEBAR ───────────────────────────────────────────────────────────────
    with st.sidebar:
        st.markdown("<div class='section-hdr'>DISPLAY</div>", unsafe_allow_html=True)
        show_months = st.slider("Import/export history (months)", 6, 48, 18, 3)
        yoy_series = st.selectbox("YoY comparison series",
            ["Total stock", "Net flow", "Imports", "Bunker sales", "Days cover"])

        st.markdown("<div class='section-hdr' style='margin-top:20px'>FORECAST ASSUMPTIONS</div>",
                    unsafe_allow_html=True)
        st.markdown(
            "<p style='font-family:IBM Plex Mono,monospace;font-size:10px;color:#3a4060'>"
            f"Imports base: 1,700 kt/mo<br>"
            f"Exports base: 610 kt/mo<br>"
            f"Bunker base: 2,500 kt/mo<br>"
            f"Crude base: 700 kt/mo<br>"
            f"Onshore target: 23,000 kbbl<br>"
            f"Float target: 1,400 kt<br>"
            f"Mean-rev speed: 15%/mo<br>"
            f"Seasonal weight: 50%</p>",
            unsafe_allow_html=True,
        )

    # ── KPI ROW ───────────────────────────────────────────────────────────────
    st.markdown("<div class='section-hdr'>LATEST ACTUAL MONTH — " +
                latest_d.strftime("%b %Y").upper() + "</div>", unsafe_allow_html=True)

    prev = actuals.iloc[-2] if len(actuals) > 1 else latest
    stock_delta = latest["total_stock_kt"] - prev["total_stock_kt"]
    nf = latest["netflow"]
    dc = latest["days_cover"]
    pct_rank = (df[~df["is_forecast"]]["total_stock_kt"] <= latest["total_stock_kt"]).mean() * 100

    c1,c2,c3,c4,c5,c6,c7 = st.columns(7)
    c1.metric("Total stock", f"{latest['total_stock_kt']:.0f} kt",
              delta=f"{stock_delta:+.0f} kt MoM")
    c2.metric("Onshore", f"{latest['onshore_kbbl']:.0f} kbbl")
    c3.metric("Floating", f"{latest['floating_kt']:.0f} kt")
    c4.metric("Net flow", f"{nf:.0f} kt",
              delta="Tightening" if nf < 0 else "Loosening",
              delta_color="inverse" if nf < 0 else "normal")
    c5.metric("Bunker sales", f"{latest['bunker_sales']:.0f} kt")
    c6.metric("Days cover", f"{dc:.1f} d",
              delta=f"{dc - stats['days_cover_mean']:+.1f}d vs avg")
    c7.metric("Stock rank", f"{pct_rank:.0f}th pct")

    st.markdown("<br>", unsafe_allow_html=True)

    # ── TABS ─────────────────────────────────────────────────────────────────
    tab1, tab2, tab3, tab4, tab5 = st.tabs([
        "Stock levels", "Net flow", "Gross flows", "Trade flows", "Year-on-year"
    ])

    with tab1:
        st.markdown("<div class='section-hdr'>TOTAL STOCK — ONSHORE + FLOATING</div>",
                    unsafe_allow_html=True)
        st.markdown(
            "<p style='font-size:11px;color:#3a4060;font-family:IBM Plex Mono,monospace'>"
            f"Amber band = historical mean ±1 std dev ({stats['stock_mean']:.0f} ± {stats['stock_std']:.0f} kt). "
            "Blue shading = forecast period. Top panel: total. Bottom: onshore/floating split.</p>",
            unsafe_allow_html=True,
        )
        st.plotly_chart(chart_total_stock(df, stats), use_container_width=True)

        st.markdown("<div class='section-hdr' style='margin-top:4px'>DAYS OF COVER</div>",
                    unsafe_allow_html=True)
        st.markdown(
            "<p style='font-size:11px;color:#3a4060;font-family:IBM Plex Mono,monospace'>"
            f"Total stock ÷ (monthly bunker sales ÷ 30). "
            f"Historical average: {stats['days_cover_mean']:.1f} days. "
            "Current: " + f"{latest['days_cover']:.1f} days.</p>",
            unsafe_allow_html=True,
        )
        st.plotly_chart(chart_days_cover(df, stats), use_container_width=True)

    with tab2:
        st.markdown("<div class='section-hdr'>NET FLOW INDICATOR + STOCK CHANGE</div>",
                    unsafe_allow_html=True)
        st.markdown(
            "<p style='font-size:11px;color:#3a4060;font-family:IBM Plex Mono,monospace'>"
            "Net flow = Net imports + Production − Bunker sales. "
            "Red = draw (tightening). Green = build (loosening). "
            "Faded bars = forecast. "
            f"Historical mean: {stats['netflow_mean']:.0f} kt/mo.</p>",
            unsafe_allow_html=True,
        )
        st.plotly_chart(chart_netflow(df, stats), use_container_width=True)

    with tab3:
        st.markdown("<div class='section-hdr'>GROSS TRADE FLOWS</div>",
                    unsafe_allow_html=True)
        st.markdown(
            "<p style='font-size:11px;color:#3a4060;font-family:IBM Plex Mono,monospace'>"
            "Top: imports (green), bunker sales (amber), exports (red), imports−bunker gap (grey fill). "
            "Bottom: crude imports + refinery production stacked. "
            "Dashed = forecast continuation.</p>",
            unsafe_allow_html=True,
        )
        st.plotly_chart(chart_gross_flows(df), use_container_width=True)

    with tab4:
        st.markdown("<div class='section-hdr'>IMPORT ORIGINS — TOP 8 COUNTRIES</div>",
                    unsafe_allow_html=True)
        st.markdown(
            f"<p style='font-size:11px;color:#3a4060;font-family:IBM Plex Mono,monospace'>"
            f"Last {show_months} months. Adjust in sidebar.</p>",
            unsafe_allow_html=True,
        )
        st.plotly_chart(chart_import_origins(df_imp, show_months), use_container_width=True)

        st.markdown("<div class='section-hdr' style='margin-top:6px'>EXPORT DESTINATIONS — TOP 8</div>",
                    unsafe_allow_html=True)
        st.plotly_chart(chart_export_destinations(df_exp, show_months), use_container_width=True)

    with tab5:
        st.markdown("<div class='section-hdr'>YEAR-ON-YEAR COMPARISON</div>",
                    unsafe_allow_html=True)
        st.markdown(
            "<p style='font-size:11px;color:#3a4060;font-family:IBM Plex Mono,monospace'>"
            "Each line = one calendar year. Current year is brightest. "
            "Shows seasonal patterns and structural shifts.</p>",
            unsafe_allow_html=True,
        )
        yoy_map = {
            "Total stock":  ("total_stock_kt",  "Total stock (kt)"),
            "Net flow":     ("netflow",          "Net flow (kt)"),
            "Imports":      ("imports",          "Imports (kt)"),
            "Bunker sales": ("bunker_sales",     "Bunker sales (kt)"),
            "Days cover":   ("days_cover",       "Days of cover"),
        }
        col, label = yoy_map[yoy_series]
        st.plotly_chart(chart_yoy(df, col, label, C["stock"]), use_container_width=True)

        # Summary table
        st.markdown("<div class='section-hdr' style='margin-top:6px'>BALANCE DATA TABLE</div>",
                    unsafe_allow_html=True)
        tbl = df[["date","is_forecast","onshore_kbbl","onshore_kt","floating_kt",
                   "total_stock_kt","stock_change","imports","exports",
                   "net_imports","bunker_sales","crude_imports","production",
                   "netflow","days_cover"]].copy()
        tbl["date"] = tbl["date"].dt.strftime("%b %Y")
        tbl["type"] = tbl["is_forecast"].map({True: "Forecast", False: "Actual"})
        tbl = tbl.drop(columns="is_forecast").rename(columns={
            "date":          "Month",
            "type":          "Type",
            "onshore_kbbl":  "Onshore (kbbl)",
            "onshore_kt":    "Onshore (kt)",
            "floating_kt":   "Floating (kt)",
            "total_stock_kt":"Total stock (kt)",
            "stock_change":  "Stock Δ (kt)",
            "imports":       "Imports (kt)",
            "exports":       "Exports (kt)",
            "net_imports":   "Net imports (kt)",
            "bunker_sales":  "Bunker sales (kt)",
            "crude_imports": "Crude imports (kt)",
            "production":    "Production (kt)",
            "netflow":       "Net flow (kt)",
            "days_cover":    "Days cover",
        })
        st.dataframe(
            tbl.sort_values("Month", ascending=False).style.format({
                c: "{:.0f}" for c in tbl.columns
                if c not in ("Month","Type","Days cover")
            } | {"Days cover": "{:.1f}"}),
            use_container_width=True, hide_index=True, height=450,
        )


if __name__ == "__main__":
    main()
